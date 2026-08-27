import pytest
from fastapi.testclient import TestClient

from app.course_registry import (
    CourseRegistryFormatError,
    build_registry_template,
    is_unresolved_code,
    parse_registry_upload,
)
from app.main import app

HEADERS = {"X-Session-Id": "test-course-registry-entries"}


def _seed_session(client: TestClient) -> None:
    resp = client.post("/api/state/restore", json={}, headers=HEADERS)
    assert resp.status_code == 200


def test_parse_registry_upload_csv_basic():
    body = b"abbreviation,course_name\nEAB,Economic Analysis for Business\nABA,Applied Business Analytics\n"
    registry, specializations, collapsed = parse_registry_upload(body, "template.csv")

    assert registry == {"EAB": "Economic Analysis for Business", "ABA": "Applied Business Analytics"}
    assert specializations == {}
    assert collapsed == []


def test_parse_registry_upload_skips_blank_and_partial_rows():
    # A blank row and a "known course names" reference row (blank
    # abbreviation, name filled) must both be silently skipped, not raise.
    body = (
        b"abbreviation,course_name\n"
        b"EAB,Economic Analysis for Business\n"
        b",\n"
        b",Financial Management\n"
    )
    registry, specializations, collapsed = parse_registry_upload(body, "template.csv")

    assert registry == {"EAB": "Economic Analysis for Business"}
    assert specializations == {}
    assert collapsed == []


def test_parse_registry_upload_case_collision_last_write_wins():
    body = (
        b"abbreviation,course_name\n"
        b"eab,Wrong Name\n"
        b"EAB,Economic Analysis for Business\n"
    )
    registry, specializations, collapsed = parse_registry_upload(body, "template.csv")

    assert registry == {"EAB": "Economic Analysis for Business"}, "last occurrence should win"
    assert collapsed == ["EAB"], "should name the colliding key, not just count it"


def test_parse_registry_upload_backward_compatible_with_old_two_column_csv():
    # Files saved before the specialization column existed must still parse.
    body = b"abbreviation,course_name\nEAB,Economic Analysis for Business\n"
    registry, specializations, collapsed = parse_registry_upload(body, "old_template.csv")

    assert registry == {"EAB": "Economic Analysis for Business"}
    assert specializations == {}
    assert collapsed == []


def test_parse_registry_upload_reads_specialization_column():
    body = (
        b"abbreviation,course_name,specialization\n"
        b"OSCSD,Operations Elective,Operations and Supply Chain\n"
        b"EAB,Economic Analysis for Business,\n"
    )
    registry, specializations, collapsed = parse_registry_upload(body, "template.csv")

    assert registry == {
        "OSCSD": "Operations Elective",
        "EAB": "Economic Analysis for Business",
    }
    # A blank specialization cell is omitted entirely, not recorded as "" --
    # see parse_registry_upload's docstring on why this must not be treated
    # as an explicit clear during a bulk (possibly partial) re-upload.
    assert specializations == {"OSCSD": "Operations and Supply Chain"}
    assert collapsed == []


def test_parse_registry_upload_rejects_unknown_extension():
    with pytest.raises(CourseRegistryFormatError):
        parse_registry_upload(b"whatever", "notes.txt")


def test_parse_registry_upload_rejects_empty_result():
    body = b"abbreviation,course_name\n,\n"
    with pytest.raises(CourseRegistryFormatError):
        parse_registry_upload(body, "template.csv")


def test_is_unresolved_code():
    assert is_unresolved_code("EAB")
    assert is_unresolved_code("A&B")
    assert not is_unresolved_code("Economic Analysis for Business")
    assert not is_unresolved_code("Applied Business Analytics")


def test_build_registry_template_csv_prefills_unresolved_codes_not_outline_names():
    content, filename, media_type = build_registry_template(
        unresolved_codes=["EAB", "ABA"],
        known_course_names=["Financial Management"],
        fmt="csv",
    )
    text = content.decode("utf-8")

    assert filename.endswith(".csv")
    assert media_type == "text/csv"
    assert "EAB" in text and "ABA" in text
    # The reference block lists known names but must not pair them with a code.
    assert "Financial Management" in text
    lines = [line for line in text.splitlines() if line.strip()]
    financial_line = next(line for line in lines if "Financial Management" in line)
    assert financial_line.startswith(","), "known course names must have a blank abbreviation column"


def test_build_registry_template_falls_back_to_examples_when_nothing_unresolved():
    content, _, _ = build_registry_template(unresolved_codes=[], known_course_names=[], fmt="csv")
    text = content.decode("utf-8")

    assert "EAB" in text  # literal example row, not derived from real data
    assert "Economic Analysis for Business" in text


def test_build_registry_template_xlsx_round_trips_through_parse():
    # A user fills in the name column after downloading the xlsx template,
    # then re-uploads it -- the second sheet (known course names) must be
    # ignored, and the filled-in row must parse back correctly.
    from openpyxl import load_workbook
    import io

    content, filename, media_type = build_registry_template(
        unresolved_codes=["EAB"], known_course_names=["Financial Management"], fmt="xlsx"
    )
    assert filename.endswith(".xlsx")
    assert "spreadsheetml" in media_type

    workbook = load_workbook(io.BytesIO(content))
    assert workbook.sheetnames == ["Course Registry", "Known course names"]
    sheet = workbook["Course Registry"]
    sheet["B2"] = "Economic Analysis for Business"  # simulate the user filling in row 2 (EAB)
    out = io.BytesIO()
    workbook.save(out)

    registry, specializations, collapsed = parse_registry_upload(out.getvalue(), "filled.xlsx")
    assert registry == {"EAB": "Economic Analysis for Business"}
    assert specializations == {}, "leaving column C blank must not invent a specialization"
    assert collapsed == []


def test_build_registry_template_xlsx_round_trips_specialization_column():
    from openpyxl import load_workbook
    import io

    content, _, _ = build_registry_template(
        unresolved_codes=["OSCSD"], known_course_names=[], fmt="xlsx"
    )
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook["Course Registry"]
    sheet["B2"] = "Operations Elective"
    sheet["C2"] = "Operations and Supply Chain"
    out = io.BytesIO()
    workbook.save(out)

    registry, specializations, collapsed = parse_registry_upload(out.getvalue(), "filled.xlsx")
    assert registry == {"OSCSD": "Operations Elective"}
    assert specializations == {"OSCSD": "Operations and Supply Chain"}
    assert collapsed == []


def test_upsert_registry_entry_adds_and_edits():
    client = TestClient(app)
    _seed_session(client)

    add = client.put("/api/course-registry/eab", json={"course_name": "Economic Analysis"}, headers=HEADERS)
    assert add.status_code == 200
    assert add.json()["abbreviation"] == "EAB", "should normalize to uppercase like the bulk upload does"

    state = client.get("/api/state", headers=HEADERS).json()
    assert state["calendar"]["course_registry"] == {"EAB": "Economic Analysis"}

    edit = client.put("/api/course-registry/EAB", json={"course_name": "Economic Analysis for Business"}, headers=HEADERS)
    assert edit.status_code == 200

    state = client.get("/api/state", headers=HEADERS).json()
    assert state["calendar"]["course_registry"] == {"EAB": "Economic Analysis for Business"}


def test_upsert_registry_entry_rejects_blank_name():
    client = TestClient(app)
    _seed_session(client)

    resp = client.put("/api/course-registry/EAB", json={"course_name": "   "}, headers=HEADERS)
    assert resp.status_code == 400

    state = client.get("/api/state", headers=HEADERS).json()
    assert state["calendar"]["course_registry"] == {}


def test_remove_registry_entry():
    client = TestClient(app)
    _seed_session(client)
    client.put("/api/course-registry/EAB", json={"course_name": "Economic Analysis"}, headers=HEADERS)
    client.put("/api/course-registry/ABA", json={"course_name": "Applied Business Analytics"}, headers=HEADERS)

    resp = client.delete("/api/course-registry/eab", headers=HEADERS)
    assert resp.status_code == 200

    state = client.get("/api/state", headers=HEADERS).json()
    assert state["calendar"]["course_registry"] == {"ABA": "Applied Business Analytics"}


def test_remove_registry_entry_missing_404s():
    client = TestClient(app)
    _seed_session(client)

    resp = client.delete("/api/course-registry/NOPE", headers=HEADERS)
    assert resp.status_code == 404


def test_upsert_registry_entry_sets_and_clears_specialization():
    client = TestClient(app)
    _seed_session(client)

    add = client.put(
        "/api/course-registry/OSCSD",
        json={"course_name": "Operations Elective", "specialization": "Operations and Supply Chain"},
        headers=HEADERS,
    )
    assert add.status_code == 200

    state = client.get("/api/state", headers=HEADERS).json()
    assert state["calendar"]["course_specializations"] == {"OSCSD": "Operations and Supply Chain"}

    # Blank specialization on a re-submit is an explicit clear (the opposite
    # of the bulk-upload path below) -- a UI dropdown always resubmits its
    # full current value, so there is no separate "leave unchanged" state.
    cleared = client.put(
        "/api/course-registry/OSCSD",
        json={"course_name": "Operations Elective", "specialization": ""},
        headers=HEADERS,
    )
    assert cleared.status_code == 200

    state = client.get("/api/state", headers=HEADERS).json()
    assert state["calendar"]["course_specializations"] == {}
    assert state["calendar"]["course_registry"] == {"OSCSD": "Operations Elective"}, (
        "clearing the specialization must not touch the course name"
    )


def test_upsert_registry_entry_specialization_normalizes_key_case_insensitively():
    # A lower-case abbreviation entered through the UI must join the same
    # normalized key ("SDT") that a parsed TimetableEntry.course_code always
    # uses (course_code is extracted via an all-caps-only regex, so it is
    # never lower-case) -- this is the registry side of that normalization
    # contract.
    client = TestClient(app)
    _seed_session(client)

    resp = client.put(
        "/api/course-registry/sdt",
        json={"course_name": "Strategic Decision Tools", "specialization": "Marketing"},
        headers=HEADERS,
    )
    assert resp.status_code == 200

    state = client.get("/api/state", headers=HEADERS).json()
    assert state["calendar"]["course_specializations"] == {"SDT": "Marketing"}


def test_remove_registry_entry_also_removes_specialization():
    client = TestClient(app)
    _seed_session(client)
    client.put(
        "/api/course-registry/EAB",
        json={"course_name": "Economic Analysis", "specialization": "Finance"},
        headers=HEADERS,
    )

    resp = client.delete("/api/course-registry/eab", headers=HEADERS)
    assert resp.status_code == 200

    state = client.get("/api/state", headers=HEADERS).json()
    assert state["calendar"]["course_registry"] == {}
    assert state["calendar"]["course_specializations"] == {}, "must not leave an orphaned tag behind"


def test_clear_registry_also_clears_specializations():
    client = TestClient(app)
    _seed_session(client)
    client.put(
        "/api/course-registry/EAB",
        json={"course_name": "Economic Analysis", "specialization": "Finance"},
        headers=HEADERS,
    )

    resp = client.delete("/api/course-registry", headers=HEADERS)
    assert resp.status_code == 200

    state = client.get("/api/state", headers=HEADERS).json()
    assert state["calendar"]["course_registry"] == {}
    assert state["calendar"]["course_specializations"] == {}


def test_bulk_upload_blank_specialization_does_not_clear_existing_tag():
    client = TestClient(app)
    _seed_session(client)
    client.put(
        "/api/course-registry/EAB",
        json={"course_name": "Economic Analysis", "specialization": "Finance"},
        headers=HEADERS,
    )

    body = b"abbreviation,course_name,specialization\nEAB,Economic Analysis for Business,\n"
    resp = client.post(
        "/api/course-registry",
        files={"file": ("update.csv", body, "text/csv")},
        headers=HEADERS,
    )
    assert resp.status_code == 200

    state = client.get("/api/state", headers=HEADERS).json()
    assert state["calendar"]["course_registry"] == {"EAB": "Economic Analysis for Business"}, (
        "the name column should still update"
    )
    assert state["calendar"]["course_specializations"] == {"EAB": "Finance"}, (
        "a blank specialization cell on a bulk re-upload must leave an existing tag untouched"
    )


def test_course_specialization_survives_state_restore():
    client = TestClient(app)
    _seed_session(client)
    client.put(
        "/api/course-registry/EAB",
        json={"course_name": "Economic Analysis", "specialization": "Finance"},
        headers=HEADERS,
    )

    exported = client.get("/api/export", headers=HEADERS).json()
    assert exported["calendar"]["course_specializations"] == {"EAB": "Finance"}

    restore = client.post("/api/state/restore", json=exported, headers=HEADERS)
    assert restore.status_code == 200

    state = client.get("/api/state", headers=HEADERS).json()
    assert state["calendar"]["course_specializations"] == {"EAB": "Finance"}, (
        "must survive a full serialize/restore round trip, not just an in-memory update"
    )
