from __future__ import annotations

import csv
import io
import re
from typing import Literal

from openpyxl import Workbook, load_workbook

_HEADER = ("abbreviation", "course_name")
MAX_UPLOAD_BYTES = 1 * 1024 * 1024
MAX_ROWS = 500

# Mirrors timetable_grid_parser._LEADING_CODE_RE's shape exactly (leading
# all-caps run of letters/&). A course_guess that still fullmatches this is
# guaranteed to be an as-extracted raw code, never a resolved name: both
# resolution paths (this registry, and orchestrator.resolve_confirmation's
# free-text answers) replace course_guess with human text that is never
# purely all-caps letters/& with nothing else.
_BARE_CODE_RE = re.compile(r"[A-Z][A-Z&]+")


def is_unresolved_code(value: str) -> bool:
    return bool(_BARE_CODE_RE.fullmatch(value))


def normalize_abbreviation(value: str) -> str:
    """The one normalization rule for registry keys, shared by the bulk
    CSV/XLSX upload and the single-entry add/edit/remove endpoints so the
    two paths can never silently disagree on what counts as the same key."""
    return value.strip().upper()


class CourseRegistryFormatError(ValueError):
    pass


def _rows_from_csv(body: bytes) -> list[list[str]]:
    text = body.decode("utf-8-sig")
    return list(csv.reader(io.StringIO(text)))


def _rows_from_xlsx(body: bytes) -> list[list[str]]:
    # Only the first sheet is data -- a second "Known course names" sheet
    # (see build_registry_template) is reference-only and deliberately
    # never read back.
    workbook = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    return [["" if cell is None else str(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]


def parse_registry_upload(body: bytes, filename: str) -> tuple[dict[str, str], list[str]]:
    """Parses an uploaded course-registry CSV/XLSX into an ABBR -> name dict.

    Returns (registry, collapsed_keys): collapsed_keys names any normalized
    abbreviation that appeared more than once in the upload (last row wins),
    so the caller can report exactly what happened rather than a bare count.

    Row 0 is always treated as a header and skipped. Any other row is only
    counted if BOTH columns are non-blank -- this is also what safely skips
    a template's blank-abbreviation "known course names" reference rows
    without needing a separate marker to detect.
    """
    if len(body) > MAX_UPLOAD_BYTES:
        raise CourseRegistryFormatError("File exceeds the 1 MB course registry upload limit")

    lower = filename.lower()
    if lower.endswith(".xlsx"):
        rows = _rows_from_xlsx(body)
    elif lower.endswith(".csv"):
        rows = _rows_from_csv(body)
    else:
        raise CourseRegistryFormatError("Only .csv or .xlsx files are accepted")

    data_rows = rows[1:][:MAX_ROWS]

    registry: dict[str, str] = {}
    seen_count: dict[str, int] = {}
    for row in data_rows:
        abbr = (row[0] if len(row) > 0 else "").strip()
        name = (row[1] if len(row) > 1 else "").strip()
        if not abbr or not name:
            continue
        key = normalize_abbreviation(abbr)
        seen_count[key] = seen_count.get(key, 0) + 1
        registry[key] = name  # last occurrence wins, matches the upsert rule in main.py

    if not registry:
        raise CourseRegistryFormatError(
            "No abbreviation/course name rows found -- check both columns are filled in"
        )

    collapsed_keys = sorted(k for k, count in seen_count.items() if count > 1)
    return registry, collapsed_keys


def _example_rows() -> list[tuple[str, str]]:
    return [("EAB", "Economic Analysis for Business"), ("ABA", "Applied Business Analytics")]


def build_registry_template(
    unresolved_codes: list[str],
    known_course_names: list[str],
    fmt: Literal["csv", "xlsx"],
) -> tuple[bytes, str, str]:
    """Builds a downloadable template.

    The abbreviation column is prefilled from `unresolved_codes` -- the
    distinct timetable course_guess codes still causing identity
    confirmation questions -- NOT from course-outline names
    (session.calendar.courses), which is a different vocabulary the
    registry exists to bridge, not reuse; prefilling from it would leave
    the user nothing to anchor the mapping to. Falls back to literal
    example rows when nothing is unresolved yet (fresh session, no
    timetable uploaded). `known_course_names` is listed separately, purely
    as a spelling reference -- never auto-paired with a code, since that
    pairing isn't known anywhere in the system.
    """
    primary_rows: list[tuple[str, str]] = (
        [(code, "") for code in sorted(unresolved_codes)] if unresolved_codes else _example_rows()
    )

    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(_HEADER)
        writer.writerows(primary_rows)
        if known_course_names:
            writer.writerow(["", ""])
            writer.writerow(
                ["", "--- Known course names (for reference; copy the spelling into a row above) ---"]
            )
            for name in sorted(known_course_names):
                writer.writerow(["", name])
        return buf.getvalue().encode("utf-8"), "course_registry_template.csv", "text/csv"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Course Registry"
    sheet.append(_HEADER)
    for row in primary_rows:
        sheet.append(row)
    if known_course_names:
        ref_sheet = workbook.create_sheet("Known course names")
        ref_sheet.append(("course_name",))
        for name in sorted(known_course_names):
            ref_sheet.append((name,))
    out = io.BytesIO()
    workbook.save(out)
    return (
        out.getvalue(),
        "course_registry_template.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
